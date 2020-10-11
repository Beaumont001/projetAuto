from askFromData import *
import openpyxl
import tkinter as tk
from tkinter.filedialog import askopenfilename

def openExcelFile(filePath):
    return openpyxl.load_workbook(filename = filePath, data_only = True, read_only = True, keep_vba = False)

def selectFile(msg):
    """Display a popup to ask the user to select a fill to use

    Args:
        the message to be displayed in the popup

    Returns:
        the path of the file name selected  

    """
    
    popup = tk.Tk()
    popup.title("Faire un choix de fichier")
    label = Label(popup, text=msg)
    label.pack(side="top", fill="x", pady=10)
    B1 = Button(popup, text="OK", command = popup.destroy)
    B1.pack()
    popup.mainloop()

    root = tk.Tk()
    root.withdraw()
    pathFile = askopenfilename(parent=root)
    root.destroy()

    return pathFile



def getSuiviSheet(workbook,worksheetName):
        '''Looks for the "Suivi [MONTH]" worksheet in the given workbook.

        Args:
            - workbook: The workbook to work on
        
        Returns:
            The "Suivi [MONTH]" worksheet in the workbook, or, if the worksheet
            was not found, None.
        '''
        worksheetNames = workbook.sheetnames

        for worksheetName in worksheetNames:
            return workbook[worksheetName]


def loadColumnNames(worksheet, headerRowNumber, columnLabelDict, headerColumnNumber = 0):
    '''This function identifies the columns in the given worksheet row.

    Args:
        worksheet: The worksheet to search the columns in.
        headerRowNumber: The number of the header row (1-based).
        headerColumnNumber: The number of the header column (0-based).
        columnLabelList: A list of the labels of the columns.
    
    Returns:
        A dictionary, that contains 2 entries:
            - FOUND: A dictionary of columns, associating the alias of the
                column to the number of the column (zero-based).
            - MISSING: A list of the aliases of the columns that have not
                been found.
    '''

    found = {}

    # Get the header row
    iterator = worksheet.iter_rows(min_row = headerRowNumber)
    row = next(iterator)

    # This counter holds the current column
    currentColumn = headerColumnNumber

    # Build reversed dict
    reversedDict = {}

    for k in columnLabelDict:
        reversedDict[columnLabelDict[k]] = k

    # While all the columns have not been found and there are column headers remaining
    while not all(x in found for x in columnLabelDict) and currentColumn < len(row):
        cellValue = row[currentColumn].value.upper()
        if cellValue in reversedDict:
            alias = reversedDict[cellValue]

            if alias in found:
                raise Exception(cellValue + ': la colonne a été trouvée au moins deux fois.')
            else:
                found[alias] = currentColumn
                #LOGGER.log(Logger.LOGLEVEL_INFO, 'Colonne \'' + cellValue.replace('\n', ' ') + '\' trouvée => ' + getColumnNumberAsString(currentColumn) + ' (' + str(currentColumn) + ')')

        currentColumn += 1

    return {
        'FOUND': found,
        'MISSING': [columnAlias for columnAlias in columnLabelDict if columnAlias not in found]
    }

def readData(worksheet,headerRow,config,dictToFill):
    data = []
    
    for worksheetRow in worksheet.iter_rows(min_row=int(headerRow) + 1):
        dictToBuild = {}
        leave = False
        for key in dictToFill['FOUND']:
            if worksheetRow[dictToFill['FOUND'][key]].value == None:
                leave = True
                break
            dictToBuild[key.lower()] = worksheetRow[dictToFill['FOUND'][key]].value
        if leave == True:
            continue
        data.append(dictToBuild)

    return data

dictOfDt = {}
for i in listOfD:
    dictOfDt[i.replace(' ','_')] = i
    
#pathFile = selectFile("boubou")
#workbook = openExcelFile(pathFile)
#a = loadColumnNames(workbook.active,4,dictOfDt,26)
#data = readData(workbook.active,4,"tes",a)
#print(len(data))